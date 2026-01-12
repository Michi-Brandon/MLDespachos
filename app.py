import asyncio
import os
import socket
import subprocess
import sys
import threading
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

try:
    from playwright.async_api import TimeoutError as PlaywrightTimeoutError
    from playwright.async_api import async_playwright
except ImportError:  # Playwright no instalado
    async_playwright = None  # type: ignore[assignment]
    PlaywrightTimeoutError = Exception  # type: ignore[assignment]

try:
    from openpyxl import load_workbook
except ImportError:  # openpyxl no instalado
    load_workbook = None  # type: ignore[assignment]


LISTING_URL = "https://www.mercadolibre.cl/ventas/omni/listado"
DETAIL_URL_TEMPLATE = "https://www.mercadolibre.cl/ventas/{code}/detalle"
LOGIN_URL = "https://www.mercadolibre.cl/ventas/omni/listado"

# Perfil dedicado para el login (con cookies)
def _base_dir() -> Path:
    # Si es un binario PyInstaller, usamos la carpeta del .exe para persistir cookies.
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys.executable).parent
    return Path(__file__).parent

BASE_DIR = _base_dir()
AUTOMATION_PROFILE_DIR = BASE_DIR / "ml_profile"

# Para apertura manual del listado con tu Chrome normal
DEFAULT_PROFILE_NAME = "Default"
DEFAULT_USER_DATA_DIR = Path(os.path.expandvars(r"%LocalAppData%\Google\Chrome\User Data"))

REMOTE_DEBUG_PORT: int | None = None
CHROME_PROCESS: subprocess.Popen | None = None


def find_chrome_executable() -> str | None:
    """
    Busca chrome.exe en rutas habituales de Windows.
    """
    candidates = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            return path
    return None


def build_chrome_command(chrome_exe: str, url: str) -> list[str]:
    """
    Prepara el comando para abrir Chrome con el perfil por defecto y la URL objetivo.
    """
    return [
        chrome_exe,
        f"--user-data-dir={DEFAULT_USER_DATA_DIR}",
        f"--profile-directory={DEFAULT_PROFILE_NAME}",
        "--new-window",
        url,
    ]


def open_with_url(url: str) -> None:
    chrome_exe = find_chrome_executable()
    if not chrome_exe:
        messagebox.showerror(
            "Chrome no encontrado",
            "No se encontro Google Chrome en las rutas tipicas. "
            "Instala Chrome o ajusta el codigo con la ruta correcta.",
        )
        return

    try:
        subprocess.Popen(
            build_chrome_command(chrome_exe, url),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as exc:  # pragma: no cover
        messagebox.showerror("Error al abrir", f"No se pudo abrir Chrome:\n{exc}")


def open_listing() -> None:
    open_with_url(LISTING_URL)


def open_login() -> None:
    print("[login] Abriendo ventana para iniciar sesion con perfil ml_profile...")
    threading.Thread(target=start_login_browser, daemon=True).start()


def open_detail(code: str) -> None:
    clean_code = code.strip()
    if not clean_code:
        messagebox.showwarning("Codigo requerido", "Ingresa un codigo de venta.")
        return

    url = DETAIL_URL_TEMPLATE.format(code=clean_code)
    print(f"[{clean_code}] Abriendo detalle y extrayendo Envíos...")

    threading.Thread(
        target=lambda: asyncio.run(open_detail_and_extract(clean_code, url)),
        daemon=True,
    ).start()


def select_and_process_excel(
    on_progress=None,
    on_status=None,
    on_finish=None,
    cancel_event: threading.Event | None = None,
) -> None:
    file_path = filedialog.askopenfilename(
        title="Seleccionar Excel",
        filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
    )
    if not file_path:
        if on_finish:
            on_finish(cancelled=False, started=False)
        return

    if on_status:
        on_status(f"Procesando: {Path(file_path).name}")
    print(f"[excel] Archivo seleccionado: {file_path}")

    def runner() -> None:
        cancelled = False
        try:
            cancelled = asyncio.run(
                process_excel(
                    file_path,
                    on_progress=on_progress,
                    on_status=on_status,
                    cancel_event=cancel_event,
                )
            )
        except Exception as exc:  # pragma: no cover - log unexpected thread error
            print(f"[excel] Error no controlado: {exc}")
            if on_status:
                on_status(f"Error: {exc}")
        finally:
            if on_finish:
                on_finish(cancelled, started=True)

    threading.Thread(target=runner, daemon=True).start()


def center_window(win: tk.Tk, width: int = 520, height: int = 380) -> None:
    win.update_idletasks()
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 3) - (height / 2))
    win.geometry(f"{width}x{height}+{x}+{y}")


def main() -> None:
    root = tk.Tk()
    root.title("Despachos ML")
    root.resizable(False, False)
    center_window(root)

    root.configure(bg="#f2f2f2")

    title = tk.Label(
        root,
        text="Abrir ventas de Mercado Libre",
        font=("Segoe UI", 12, "bold"),
        bg="#f2f2f2",
    )
    title.pack(pady=(20, 10))

    description = tk.Label(
        root,
        text="1) Inicia sesion (boton abajo). 2) Procesa tu Excel de MercadoLibre.",
        font=("Segoe UI", 9),
        bg="#f2f2f2",
    )
    description.pack(pady=(0, 16))

    login_button = tk.Button(
        root,
        text="Login y guardar cookies (perfil ML)",
        font=("Segoe UI", 10, "bold"),
        bg="#03a9f4",
        fg="#fff",
        activebackground="#039be5",
        activeforeground="#fff",
        relief=tk.FLAT,
        padx=16,
        pady=8,
        command=open_login,
        cursor="hand2",
    )
    login_button.pack(pady=(0, 14))

    progress_var = tk.StringVar(value="Progreso: 0/0")
    status_var = tk.StringVar(value="Listo para procesar.")

    progress_frame = tk.Frame(root, bg="#f2f2f2")
    progress_frame.pack(pady=(4, 6), fill="x")

    progress_label = tk.Label(
        progress_frame,
        textvariable=progress_var,
        font=("Segoe UI", 9, "bold"),
        bg="#f2f2f2",
        anchor="w",
    )
    progress_label.pack(anchor="w", padx=12)

    progress_bar = ttk.Progressbar(
        progress_frame,
        orient="horizontal",
        mode="determinate",
        length=360,
    )
    progress_bar.pack(fill="x", padx=12, pady=(2, 6))

    status_label = tk.Label(
        progress_frame,
        textvariable=status_var,
        font=("Segoe UI", 9),
        bg="#f2f2f2",
        anchor="w",
        wraplength=460,
        justify="left",
    )
    status_label.pack(anchor="w", padx=12, pady=(0, 8))

    process_button: tk.Button
    cancel_button: tk.Button
    current_cancel_event: threading.Event | None = None

    def set_processing_state(is_running: bool) -> None:
        if is_running:
            process_button.config(state=tk.DISABLED)
            cancel_button.config(state=tk.NORMAL)
        else:
            process_button.config(state=tk.NORMAL)
            cancel_button.config(state=tk.DISABLED)

    def update_progress(processed: int, total: int) -> None:
        def _update() -> None:
            total_for_bar = max(total, 1)
            progress_bar.config(maximum=total_for_bar)
            progress_bar["value"] = processed
            progress_var.set(f"Progreso: {processed}/{total}")

        root.after(0, _update)

    def update_status(text: str) -> None:
        root.after(0, lambda: status_var.set(text))

    def finish_processing(cancelled: bool, started: bool = True) -> None:
        def _finish() -> None:
            set_processing_state(False)
            if not started:
                progress_var.set("Progreso: 0/0")
                progress_bar["value"] = 0
                status_var.set("Listo para procesar.")
                return
            if cancelled and not status_var.get():
                status_var.set("Proceso cancelado.")

        root.after(0, _finish)

    def cancel_processing() -> None:
        nonlocal current_cancel_event
        if current_cancel_event and not current_cancel_event.is_set():
            current_cancel_event.set()
            update_status("Cancelando proceso...")

    def start_excel_processing() -> None:
        nonlocal current_cancel_event
        current_cancel_event = threading.Event()
        set_processing_state(True)
        update_progress(0, 0)
        update_status("Selecciona un archivo de Excel...")
        select_and_process_excel(
            on_progress=update_progress,
            on_status=update_status,
            on_finish=finish_processing,
            cancel_event=current_cancel_event,
        )

    process_button = tk.Button(
        root,
        text="Procesar Excel (ML y Walmart)",
        font=("Segoe UI", 10, "bold"),
        bg="#9c27b0",
        fg="#fff",
        activebackground="#8e24aa",
        activeforeground="#fff",
        relief=tk.FLAT,
        padx=16,
        pady=8,
        command=start_excel_processing,
        cursor="hand2",
    )
    process_button.pack(pady=(0, 10))

    cancel_button = tk.Button(
        root,
        text="Cancelar proceso",
        font=("Segoe UI", 10, "bold"),
        bg="#e53935",
        fg="#fff",
        activebackground="#d32f2f",
        activeforeground="#fff",
        relief=tk.FLAT,
        padx=16,
        pady=8,
        command=cancel_processing,
        cursor="hand2",
        state=tk.DISABLED,
    )
    cancel_button.pack(pady=(0, 12))

    root.mainloop()


async def open_detail_and_extract(code: str, url: str) -> None:
    """
    Se conecta al Chrome abierto con el boton de login (puerto CDP) y lee el valor de Envíos.
    Se crea una instancia nueva de Playwright por llamada para evitar conexiones rotas.
    """
    if async_playwright is None:
        print(
            f"[{code}] Playwright no esta instalado. Ejecuta: pip install playwright && python -m playwright install"
        )
        return

    if REMOTE_DEBUG_PORT is None:
        print(f"[{code}] No hay puerto de depuracion. Pulsa primero el boton de login.")
        return

    if not wait_for_port("localhost", REMOTE_DEBUG_PORT, attempts=10, delay=0.4):
        print(f"[{code}] No se pudo alcanzar el puerto {REMOTE_DEBUG_PORT}.")
        return

    endpoint = f"http://localhost:{REMOTE_DEBUG_PORT}"
    playwright = await async_playwright().start()
    try:
        browser = await playwright.chromium.connect_over_cdp(endpoint)
        if not browser.contexts:
            print(f"[{code}] No hay contextos en Chrome. ¿Cerraste la ventana de login?")
            return

        context = browser.contexts[0]
        try:
            page = await context.new_page()
        except Exception as exc:
            print(f"[{code}] No se pudo abrir una nueva pestaña: {exc}")
            return

        page.set_default_timeout(20000)
        await page.goto(url, wait_until="domcontentloaded")

        text = await extract_amount_text(page, "Envíos", timeout_ms=2000)
        source = "Envíos"

        if text is None:
            # Bonificaciones debería resolverse rápido si no hay Envíos; usa timeout corto
            text = await extract_amount_text(page, "Bonificaciones", timeout_ms=1000)
            source = "Bonificaciones" if text else None

        if text is None:
            print(f"[{code}] No se encontraron Envíos ni Bonificaciones. Valor: $ 0")
            return

        parsed = parse_amount(text)
        if parsed is None:
            print(f"[{code}] No se pudo interpretar el valor de {source}: {text}")
            return

        if parsed < 0:
            parsed = 0

        print(f"[{code}] Envíos ({source}): {format_amount(parsed)}")
    except Exception as exc:
        print(f"[{code}] Error conectando a Chrome: {exc}")
    finally:
        try:
            await playwright.stop()
        except Exception:
            pass


async def process_excel(
    file_path: str,
    on_progress=None,
    on_status=None,
    cancel_event: threading.Event | None = None,
) -> bool:
    def notify_status(message: str) -> None:
        if on_status:
            on_status(message)

    def notify_progress(done: int, total: int) -> None:
        if on_progress:
            on_progress(done, total)

    if load_workbook is None:
        msg = "[excel] Falta openpyxl. Instala con: pip install openpyxl"
        print(msg)
        notify_status("Falta openpyxl. Instala con: pip install openpyxl")
        return False
    if async_playwright is None:
        msg = "[excel] Falta Playwright. Instala con: pip install playwright && python -m playwright install"
        print(msg)
        notify_status("Falta Playwright. Instala con: pip install playwright && python -m playwright install")
        return False
    if REMOTE_DEBUG_PORT is None:
        print("[excel] No hay puerto de depuracion. Pulsa el boton de login primero.")
        notify_status("No hay puerto de depuracion. Pulsa el boton de login primero.")
        return False
    if not wait_for_port("localhost", REMOTE_DEBUG_PORT, attempts=10, delay=0.4):
        print(f"[excel] No se pudo alcanzar el puerto {REMOTE_DEBUG_PORT}.")
        notify_status(f"No se pudo alcanzar el puerto {REMOTE_DEBUG_PORT}.")
        return False

    notify_status("Abriendo Excel...")
    try:
        wb = load_workbook(file_path)
    except Exception as exc:
        print(f"[excel] No se pudo abrir el archivo: {exc}")
        notify_status(f"No se pudo abrir el archivo: {exc}")
        return False

    if "Reporte" not in wb.sheetnames:
        print("[excel] No se encontro la hoja 'Reporte'.")
        notify_status("No se encontro la hoja 'Reporte' en el Excel.")
        return False

    ws = wb["Reporte"]
    max_row = ws.max_row

    last_data_col = 0
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            if cell.column > last_data_col:
                last_data_col = cell.column

    if last_data_col < 3:
        notify_status("No se pudo detectar la ultima columna con datos.")
        print("[excel] No se pudo detectar la ultima columna con datos.")
        return False

    w_col = last_data_col - 2
    x_col = last_data_col - 1
    y_col = last_data_col

    rows_to_process: list[tuple[int, str]] = []
    walmart_groups: dict[str, list[int]] = {}
    for row_idx in range(2, max_row + 1):
        channel = ws.cell(row=row_idx, column=6).value  # F
        sale_code = ws.cell(row=row_idx, column=8).value  # H
        channel_norm = str(channel).strip().lower()
        if channel_norm == "mercadolibre":
            if not sale_code:
                continue
            rows_to_process.append((row_idx, sale_code))
        elif channel_norm == "walmart":
            if not sale_code:
                continue
            code_key = str(sale_code).strip()
            if not code_key:
                continue
            walmart_groups.setdefault(code_key, []).append(row_idx)

    total_rows = len(rows_to_process)
    notify_progress(0, total_rows)
    if total_rows == 0:
        notify_status("No hay filas de MercadoLibre para procesar.")
        print("[excel] No hay filas de MercadoLibre para procesar.")

    endpoint = f"http://localhost:{REMOTE_DEBUG_PORT}"
    playwright = await async_playwright().start()
    processed_ml = 0
    processed_walmart = 0
    total_walmart_rows = sum(len(rows) for rows in walmart_groups.values())
    cancelled = False
    try:
        browser = await playwright.chromium.connect_over_cdp(endpoint)
        if not browser.contexts:
            print("[excel] No hay contextos en Chrome. ¿Cerraste la ventana de login?")
            notify_status("No hay contextos en Chrome. ¿Cerraste la ventana de login?")
            return False
        context = browser.contexts[0]

        if total_rows > 0:
            notify_status("Procesando MercadoLibre...")
            for row_idx, sale_code in rows_to_process:
                if cancel_event and cancel_event.is_set():
                    cancelled = True
                    notify_status(f"Proceso cancelado. Guardando archivo... ({processed_ml}/{total_rows})")
                    break

                url = DETAIL_URL_TEMPLATE.format(code=sale_code)
                amount = await fetch_amount_for_code(context, sale_code, url)
                if amount is None:
                    amount = 0

                ws.cell(row=row_idx, column=x_col).value = amount  # X

                w_raw = ws.cell(row=row_idx, column=w_col).value  # W
                w_val = parse_amount(w_raw)
                w_val = w_val if w_val is not None else 0
                ws.cell(row=row_idx, column=y_col).value = w_val + amount  # Y

                processed_ml += 1
                notify_progress(processed_ml, total_rows)
                print(f"[excel] Fila {row_idx} ({sale_code}) -> Envíos: {format_amount(amount)}")

        if not cancelled and total_walmart_rows > 0:
            notify_progress(0, total_walmart_rows)
            notify_status("Procesando Walmart...")
            for code_key, row_indices in walmart_groups.items():
                if cancel_event and cancel_event.is_set():
                    cancelled = True
                    notify_status(
                        f"Proceso cancelado. Guardando archivo... ({processed_walmart}/{total_walmart_rows})"
                    )
                    break

                prices: list[int] = []
                totals: list[int] = []
                for row_idx in row_indices:
                    price_val = parse_amount(ws.cell(row=row_idx, column=w_col).value)
                    total_val = parse_amount(ws.cell(row=row_idx, column=y_col).value)
                    prices.append(price_val if price_val is not None else 0)
                    totals.append(total_val if total_val is not None else 0)

                sum_prices = sum(prices)
                group_total = max(totals) if totals else 0
                diff = group_total - sum_prices
                if diff < 0:
                    diff = 0

                first_row = row_indices[0]
                for row_idx in row_indices:
                    if cancel_event and cancel_event.is_set():
                        cancelled = True
                        break
                    ws.cell(row=row_idx, column=x_col).value = 0  # Despacho
                    processed_walmart += 1
                    notify_progress(processed_walmart, total_walmart_rows)
                if cancelled:
                    notify_status(
                        f"Proceso cancelado. Guardando archivo... ({processed_walmart}/{total_walmart_rows})"
                    )
                    break
                if diff > 0:
                    ws.cell(row=first_row, column=x_col).value = diff  # Despacho

            if not cancelled:
                notify_status("Walmart terminado.")

        out_path = Path(file_path)
        output_file = out_path.with_name(f"{out_path.stem}_con_envios{out_path.suffix}")
        wb.save(output_file)
        if cancelled:
            message = (
                "Proceso cancelado. "
                f"MercadoLibre: {processed_ml}/{total_rows}. "
                f"Walmart: {processed_walmart}/{total_walmart_rows}. "
                f"Archivo: {output_file}"
            )
        else:
            message = (
                "Listo. "
                f"MercadoLibre: {processed_ml}/{total_rows}. "
                f"Walmart: {processed_walmart}/{total_walmart_rows}. "
                f"Archivo guardado en: {output_file}"
            )
        print(f"[excel] {message}")
        notify_status(message)
        return cancelled
    except Exception as exc:
        print(f"[excel] Error procesando Excel: {exc}")
        notify_status(f"Error procesando Excel: {exc}")
        return False
    finally:
        try:
            await playwright.stop()
        except Exception:
            pass


async def fetch_amount_for_code(context, code: str, url: str) -> int | None:
    try:
        page = await context.new_page()
    except Exception as exc:
        print(f"[{code}] No se pudo abrir una nueva pestaña: {exc}")
        return None

    try:
        page.set_default_timeout(20000)
        await page.goto(url, wait_until="domcontentloaded")

        text = await extract_amount_text(page, "Envíos", timeout_ms=2000)
        source = "Envíos"

        if text is None:
            text = await extract_amount_text(page, "Bonificaciones", timeout_ms=1000)
            source = "Bonificaciones" if text else None

        if text is None:
            print(f"[{code}] No se encontraron Envíos ni Bonificaciones. Valor: $ 0")
            return 0

        parsed = parse_amount(text)
        if parsed is None:
            print(f"[{code}] No se pudo interpretar el valor de {source}: {text}")
            return None

        if parsed < 0:
            parsed = 0

        print(f"[{code}] Envíos ({source}): {format_amount(parsed)}")
        return parsed
    except PlaywrightTimeoutError:
        print(f"[{code}] Timeout esperando datos. Revisa si hay login pendiente.")
        return None
    except Exception as exc:
        print(f"[{code}] Error extrayendo datos: {exc}")
        return None
    finally:
        try:
            await page.close()
        except Exception:
            pass

def find_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("localhost", 0))
        return s.getsockname()[1]


def wait_for_port(host: str, port: int, attempts: int = 15, delay: float = 0.4) -> bool:
    for _ in range(attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(delay)
            if sock.connect_ex((host, port)) == 0:
                return True
        time.sleep(delay)
    return False


def start_login_browser(start_url: str = LOGIN_URL) -> None:
    """
    Lanza Chrome normal (sin banderas de automatizacion) con remote debugging para que el usuario
    haga login y guarde cookies en ./ml_profile. Mantiene el puerto en REMOTE_DEBUG_PORT.
    """
    global REMOTE_DEBUG_PORT, CHROME_PROCESS

    chrome_exe = find_chrome_executable()
    if not chrome_exe:
        messagebox.showerror(
            "Chrome no encontrado",
            "No se encontro Google Chrome en las rutas tipicas. Ajusta la ruta en el codigo.",
        )
        return

    REMOTE_DEBUG_PORT = find_free_port()
    AUTOMATION_PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    args = [
        chrome_exe,
        f"--remote-debugging-port={REMOTE_DEBUG_PORT}",
        f"--user-data-dir={AUTOMATION_PROFILE_DIR}",
        "--profile-directory=Default",
        "--start-maximized",
        "--no-default-browser-check",
        "--no-first-run",
        start_url,
    ]

    try:
        CHROME_PROCESS = subprocess.Popen(
            args,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as exc:
        print(f"[login] No se pudo lanzar Chrome: {exc}")
        return

    if wait_for_port("localhost", REMOTE_DEBUG_PORT):
        print(
            f"[login] Chrome abierto en {start_url}. Puerto CDP {REMOTE_DEBUG_PORT}. "
            "Inicia sesion; las cookies se guardan en ./ml_profile."
        )
    else:
        print("[login] No se pudo confirmar el puerto de depuracion. Reintenta.")


async def extract_amount_text(page, title: str, timeout_ms: int = 15000) -> str | None:
    """
    Intenta obtener el texto del subtotal para una fila con el titulo dado.
    """
    try:
        row = page.locator("div.sc-account-rows__row", has_text=title).first
        await row.wait_for(state="attached", timeout=timeout_ms)
        text = await row.locator("span.sc-account-rows__row__subTotal").first.text_content()
        return text.strip() if text else None
    except PlaywrightTimeoutError:
        return None
    except Exception:
        return None


def parse_amount(text: str) -> int | None:
    """
    Convierte un texto como "$ 3.090" o "-$ 2.276" a entero (pesos).
    """
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return int(text)
    clean = text.replace("\xa0", " ").replace("$", "")
    sign = -1 if "-" in clean else 1
    digits = "".join(ch for ch in clean if ch.isdigit())
    if not digits:
        return None
    try:
        value = int(digits) * sign
    except ValueError:
        return None
    return value


def format_amount(value: int) -> str:
    """
    Devuelve una cadena tipo "$ 3.090" con separador de miles usando punto.
    """
    return "$ " + f"{value:,}".replace(",", ".")


if __name__ == "__main__":
    main()
